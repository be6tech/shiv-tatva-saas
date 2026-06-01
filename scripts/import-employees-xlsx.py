"""
Import employee roster from Employee_Master_List.xlsx (or legacy EMPLOYEE DETAILS.xlsx).

Usage:
  python scripts/import-employees-xlsx.py
  python scripts/import-employees-xlsx.py write
  python scripts/import-employees-xlsx.py write --xlsx "C:\\Users\\...\\Employee_Master_List.xlsx"
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    raise

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path(r"c:\Users\durga\Downloads\Employee_Master_List.xlsx")
LEGACY_XLSX = ROOT / "EMPLOYEE DETAILS.xlsx"

ID_RE = re.compile(r"^(STS26[A-Z]{3}\d{3})(?:\s+(.*))?$", re.IGNORECASE)
EMAIL_RE = re.compile(r"[\w.+-]+@[\w.-]+\.\w+", re.IGNORECASE)

ROLE_META = {
    "ASE": ("ASSOCIATE TRAINEE", "Learning & Development"),
    "HRM": ("HRM", "HR"),
    "BDE": ("BDE", "Business Development"),
    "BOE": ("BDE", "Business Development"),  # sheet typo BOE → same as BDE
}


def title_name(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return "Unknown"
    return " ".join(w.capitalize() for w in s.split())


def name_from_email(email: str) -> str:
    local = email.split("@", 1)[0]
    local = re.sub(r"\d+", " ", local)
    local = local.replace(".", " ").replace("_", " ")
    return title_name(local)


def role_code(employee_id: str) -> str:
    m = re.match(r"^STS26([A-Z]{3})\d{3}$", employee_id.upper())
    return m.group(1) if m else "ASE"


def designation_for(employee_id: str, name: str) -> tuple[str, str]:
    if name.strip().lower() == "admin":
        return "ADMIN", "Operations"
    code = role_code(employee_id)
    return ROLE_META.get(code, ("ASSOCIATE TRAINEE", "Learning & Development"))


def status_for(designation: str) -> str:
    d = designation.upper()
    if "HRM" in d:
        return "HR Manager"
    if d == "ADMIN":
        return "Active"
    return "Active"


def exp_for(designation: str) -> int:
    d = designation.upper()
    if "TRAINEE" in d:
        return 0
    if "BDE" in d:
        return 2
    if "HRM" in d:
        return 8
    if "EXECUTIVE" in d:
        return 1
    return 1


def skills_for(designation: str) -> list[str]:
    d = designation.upper()
    if "BDE" in d:
        return ["Business development", "Client outreach", "CRM"]
    if "HRM" in d or d == "HR":
        return ["HR operations", "Talent management", "HRMS"]
    if "TRAINEE" in d:
        return ["On-the-job training", "Fundamentals", "Team collaboration"]
    if "EXECUTIVE" in d:
        return ["Operations", "Documentation", "Coordination"]
    return ["Professional skills"]


def dept_for_legacy(des: str) -> str:
    d = (des or "").strip().upper()
    if "HRM" in d or d == "HR":
        return "HR"
    if "DEVELOPER" in d:
        return "Engineering"
    if "BDE" in d:
        return "Business Development"
    if "TRAINEE" in d:
        return "Learning & Development"
    return "Operations"


def load_legacy_master(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    out: list[dict] = []
    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        rid = str(row[0]).strip()
        name = title_name(str(row[1]))
        des = str(row[2] or "").strip()
        email = str(row[3] or "").strip()
        if not rid or not name:
            continue
        out.append(
            {
                "id": rid,
                "name": name,
                "designation": des,
                "department": dept_for_legacy(des),
                "skills": skills_for(des),
                "experienceYears": exp_for(des),
                "email": email,
                "status": status_for(des),
                "joiningDate": "2024-06-01",
                "bio": f"{name} serves as {des} at Shiv Tatva Solutions, contributing to delivery and team growth.",
                "projects": [],
                "certifications": [],
            }
        )
    return out


def parse_master_row(row) -> dict | None:
    if not row or all(c is None or str(c).strip() == "" for c in row):
        return None

    raw_id = str(row[1] or "").strip()
    m = ID_RE.match(raw_id)
    if not m:
        emails_in_id = EMAIL_RE.findall(raw_id)
        id_match = re.match(r"^(STS26[A-Z]{3}\d{3})", raw_id.upper())
        if not id_match:
            return None
        employee_id = id_match.group(1).upper()
        tail = raw_id[id_match.end() :].strip()
        extra = tail
        email_from_id = emails_in_id[0] if emails_in_id else ""
    else:
        employee_id = m.group(1).upper()
        extra = (m.group(2) or "").strip()
        email_from_id = ""

    raw_name = str(row[2] or "").strip()
    name = title_name(raw_name) if raw_name else ""
    email = str(row[3] or "").strip()

    if not email and email_from_id:
        email = email_from_id
    if not email and extra and EMAIL_RE.search(extra):
        email = EMAIL_RE.search(extra).group(0)  # type: ignore[union-attr]
    elif not name and extra and not EMAIL_RE.search(extra):
        name = title_name(extra)

    email = email.strip().lower()
    if not name and email:
        name = name_from_email(email)
    if not employee_id or not name:
        return None

    designation, department = designation_for(employee_id, name)
    skills = skills_for(designation)
    st = status_for(designation)
    ex = exp_for(designation)
    bio = f"{name} serves as {designation} at Shiv Tatva Solutions, contributing to delivery and team growth."

    return {
        "id": employee_id,
        "name": name,
        "designation": designation,
        "department": department,
        "skills": skills,
        "experienceYears": ex,
        "email": email,
        "status": st,
        "joiningDate": "2026-01-01",
        "bio": bio,
        "projects": [],
        "certifications": [],
    }


def load_master_list(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    out: list[dict] = []
    for row in rows[1:]:
        parsed = parse_master_row(row)
        if parsed:
            out.append(parsed)
    return out


def load_employees(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Employees"] if "Employees" in wb.sheetnames else wb.active
    header = [str(c or "").strip().lower() for c in next(ws.iter_rows(max_row=1, values_only=True))]
    if "employee id" in header:
        return load_master_list(ws)
    return load_legacy_master(ws)


def to_seed_record(e: dict) -> dict:
    return {
        "id": e["id"],
        "name": e["name"],
        "department": e["department"],
        "designation": e["designation"],
        "status": e["status"] if e["status"] in ("Active", "HR Manager") else "Active",
        "email": e.get("email") or "",
        "phone": "",
        "location": "Hyderabad",
        "joinedAt": e.get("joiningDate", "2026-01-01"),
        "skills": e["skills"],
        "experienceYears": e["experienceYears"],
    }


def write_employees_ts(target: Path, employees: list[dict]) -> None:
    header = '''/**
 * Employee directory for marketing `/employee-details`.
 * Imported from `Employee_Master_List.xlsx` (Employee ID, Name, Email).
 * Re-import: `python scripts/import-employees-xlsx.py write`
 */
export type WorkStatus =
  | "Active"
  | "Team Lead"
  | "Senior Developer"
  | "HR Manager"
  | "UI/UX Designer";

export type Employee = {
  id: string;
  name: string;
  designation: string;
  department: string;
  skills: string[];
  experienceYears: number;
  email: string;
  linkedin?: string;
  status: WorkStatus;
  joiningDate: string; // YYYY-MM-DD
  bio: string;
  projects: string[];
  certifications: string[];
};

export const employees: Employee[] = [
'''
    chunks: list[str] = []
    for e in employees:
        chunks.append(
            "  {\n"
            + f'    id: {json.dumps(e["id"])},\n'
            + f'    name: {json.dumps(e["name"])},\n'
            + f'    designation: {json.dumps(e["designation"])},\n'
            + f'    department: {json.dumps(e["department"])},\n'
            + f'    skills: {json.dumps(e["skills"], ensure_ascii=False)},\n'
            + f'    experienceYears: {e["experienceYears"]},\n'
            + f'    email: {json.dumps(e["email"])},\n'
            + f'    status: {json.dumps(e["status"])},\n'
            + f'    joiningDate: {json.dumps(e["joiningDate"])},\n'
            + f'    bio: {json.dumps(e["bio"], ensure_ascii=False)},\n'
            + f'    projects: {json.dumps(e["projects"])},\n'
            + f'    certifications: {json.dumps(e["certifications"])},\n'
            + "  }"
        )
    target.write_text(header + ",\n".join(chunks) + "\n];\n", encoding="utf-8")


def write_employees_seed_js(target: Path, employees: list[dict]) -> None:
    seed = [to_seed_record(e) for e in employees]
    lines = [
        "/** HRMS employee directory — synced from Employee_Master_List.xlsx (STS26… IDs). */",
        "export const employeesSeed = [",
    ]
    for e in seed:
        lines.append("  {")
        lines.append(f'    id: {json.dumps(e["id"])},')
        lines.append(f'    name: {json.dumps(e["name"])},')
        lines.append(f'    department: {json.dumps(e["department"])},')
        lines.append(f'    designation: {json.dumps(e["designation"])},')
        lines.append(f'    status: {json.dumps(e["status"])},')
        lines.append(f'    email: {json.dumps(e["email"])},')
        lines.append('    phone: "",')
        lines.append(f'    location: {json.dumps(e["location"])},')
        lines.append(f'    joinedAt: {json.dumps(e["joinedAt"])},')
        lines.append(f'    skills: {json.dumps(e["skills"], ensure_ascii=False)},')
        lines.append(f'    experienceYears: {e["experienceYears"]},')
        lines.append("  },")
    lines.append("];")
    lines.append("")
    lines.append("export const employeeShiftSeed = Object.fromEntries(")
    lines.append("  employeesSeed.map((e) => [e.id, \"morning\"])")
    lines.append(");")
    lines.append("")
    target.write_text("\n".join(lines), encoding="utf-8")


def prune_db_for_employees(db: dict, valid_ids: set[str]) -> None:
    att = db.get("attendance") or {}
    if isinstance(att, dict):
        db["attendance"] = {
            k: v for k, v in att.items() if str(v.get("employeeId", k.split(":")[-1])) in valid_ids
        }

    shifts = db.get("employeeShift") or {}
    if isinstance(shifts, dict):
        db["employeeShift"] = {k: v for k, v in shifts.items() if k in valid_ids}

    for key in ("leaveRequests", "tasks", "payslips"):
        items = db.get(key)
        if isinstance(items, list):
            db[key] = [x for x in items if x.get("employeeId") in valid_ids]

    notifs = db.get("notifications")
    if isinstance(notifs, list):
        db["notifications"] = [
            n
            for n in notifs
            if n.get("recipientId") is None or n.get("recipientId") in valid_ids
        ]


def write_employee_auth_constants(target: Path, employees: list[dict]) -> None:
    roster = [
        {"id": e["id"], "email": e["email"]}
        for e in employees
        if (e.get("email") or "").strip()
    ]
    lines = [
        "/** All roster logins use initial password `demo` until changed via forgot-password. */",
        "export const EMPLOYEE_SEED_PASSWORD_HASH =",
        f'  "{EMPLOYEE_SEED_PASSWORD_HASH}";',
        "",
        "export const EMPLOYEE_ROSTER = [",
    ]
    for r in roster:
        lines.append(f'  {{ id: {json.dumps(r["id"])}, email: {json.dumps(r["email"])} }},')
    lines.append("] as const;")
    lines.append("")
    lines.append("export function findRosterEmployee(identifier: string) {")
    lines.append("  const value = identifier.trim();")
    lines.append("  return (")
    lines.append("    EMPLOYEE_ROSTER.find(")
    lines.append("      (e) =>")
    lines.append("        e.id.toUpperCase() === value.toUpperCase() ||")
    lines.append("        e.email.toLowerCase() === value.toLowerCase()")
    lines.append("    ) ?? null")
    lines.append("  );")
    lines.append("}")
    lines.append("")
    target.write_text("\n".join(lines), encoding="utf-8")


EMPLOYEE_SEED_PASSWORD_HASH = (
    "f4298c5a7302c99350dec80533fbbc33:d188636c159647671ad1a90842ab9b2400feb8591f2dfe4e33ae83407ced02b8c11edfa847f7526d3128cc119e745d42351248104d8beb7aea4e0d0952b1cf97"
)


def sql_literal(value: str) -> str:
    """PostgreSQL string literal (single quotes — not JSON double quotes)."""
    return "'" + value.replace("'", "''") + "'"


def write_employee_users_sql(target: Path, employees: list[dict]) -> None:
    roster = [(e["id"], e["email"]) for e in employees if (e.get("email") or "").strip()]
    values = ",\n".join(
        f"  ({sql_literal(eid)}, {sql_literal(email)}, {sql_literal(EMPLOYEE_SEED_PASSWORD_HASH)})"
        for eid, email in roster
    )
    sql = f"""-- Employee portal logins (server-only via service role). Run in Supabase SQL Editor.
-- Initial password for all accounts: demo (change after first login via forgot-password)

create table if not exists public.employee_users (
  id uuid primary key default gen_random_uuid(),
  created_at timestamptz not null default now(),
  updated_at timestamptz not null default now(),
  employee_id text not null unique,
  email text not null unique,
  password_hash text not null,
  reset_token text,
  reset_token_expires_at timestamptz
);

alter table public.employee_users enable row level security;

-- Drop legacy BE1999… / demo rows (same emails are re-inserted with STS26… IDs below)
delete from public.employee_users
where employee_id like 'BE1999%'
   or employee_id = 'ST-EMP-001';

insert into public.employee_users (employee_id, email, password_hash)
values
{values}
on conflict (email) do update set
  employee_id = excluded.employee_id,
  password_hash = excluded.password_hash,
  updated_at = now();
"""
    target.write_text(sql, encoding="utf-8")


def write_db_json(target: Path, employees: list[dict]) -> None:
    seed = [to_seed_record(e) for e in employees]
    valid_ids = {e["id"] for e in seed}
    if target.exists():
        db = json.loads(target.read_text(encoding="utf-8"))
    else:
        db = {}
    prune_db_for_employees(db, valid_ids)
    db["employees"] = seed
    db["employeeShift"] = {e["id"]: "morning" for e in seed}
    target.write_text(json.dumps(db, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("command", nargs="?", default="print")
    parser.add_argument("--xlsx", type=Path, default=None)
    args = parser.parse_args()

    xlsx = args.xlsx or DEFAULT_XLSX
    if not xlsx.exists() and LEGACY_XLSX.exists():
        xlsx = LEGACY_XLSX
    if not xlsx.exists():
        print(f"Excel not found: {xlsx}", file=sys.stderr)
        sys.exit(1)

    employees = load_employees(xlsx)
    if not employees:
        print("No employees parsed.", file=sys.stderr)
        sys.exit(1)

    if args.command == "write":
        write_employees_ts(ROOT / "apps" / "web" / "src" / "data" / "employees.ts", employees)
        write_employees_seed_js(ROOT / "services" / "api-gateway" / "src" / "employees-seed.js", employees)
        write_db_json(ROOT / "services" / "api-gateway" / "data" / "db.json", employees)
        write_employee_auth_constants(ROOT / "apps" / "web" / "src" / "lib" / "employee-auth-constants.ts", employees)
        write_employee_users_sql(ROOT / "apps" / "web" / "supabase" / "employee_users.sql", employees)
        print(f"Imported {len(employees)} employees from {xlsx}")
        print("  - apps/web/src/data/employees.ts")
        print("  - services/api-gateway/src/employees-seed.js")
        print("  - services/api-gateway/data/db.json (replaced roster, pruned old HRMS refs)")
        print("  - apps/web/src/lib/employee-auth-constants.ts")
        print("  - apps/web/supabase/employee_users.sql")
    else:
        print(json.dumps(employees, indent=2))


if __name__ == "__main__":
    main()
